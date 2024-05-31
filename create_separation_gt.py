import logging
from collections import Counter, OrderedDict, defaultdict
from pathlib import Path

import imagesize
from openpyxl import Workbook

from utils.copy_utils import copy_mode
from utils.input_utils import get_file_paths, supported_image_formats


def get_arguments():
    import argparse

    parser = argparse.ArgumentParser(description="Create separation ground truth")
    io_args = parser.add_argument_group("IO")
    io_args.add_argument("-i", "--input", help="Train input folder/file", nargs="+", action="extend", type=str, required=True)
    io_args.add_argument("-o", "--output", help="Output folder", type=str)

    parser.add_argument(
        "-m",
        "--output-mode",
        help="Output mode",
        type=str,
        choices=["xlsx", "dirs"],
        default="xlsx",
    )

    args = parser.parse_args()
    return args


def get_size_match(image_size1, image_size2, margin):
    min_multiplier = 1.0 - margin
    max_multiplier = 1.0 + margin
    border_multiplier = 0.01
    image1_width = image_size1[0]
    image2_width = image_size2[0]
    image1_height = image_size1[1]
    image2_height = image_size2[1]
    similar_height_width = (
        image1_width * min_multiplier < image2_width < image1_width * max_multiplier
        and image1_height * min_multiplier < image2_height < image1_height * max_multiplier
    )
    if similar_height_width:
        return True
    similar_height_half_width = (
        image1_width * (1 - border_multiplier) * min_multiplier
        < image2_width / 2
        < image1_width * (1 - border_multiplier) * max_multiplier
        and image1_height * min_multiplier < image2_height < image1_height * max_multiplier
    )
    if similar_height_half_width:
        return True
    similar_height_double_width = (
        image1_width * min_multiplier < image2_width * (2 - border_multiplier) < image1_width * max_multiplier
        and image1_height * min_multiplier < image2_height < image1_height * max_multiplier
    )
    if similar_height_double_width:
        return True
    return False


def main(args):
    logger = logging.getLogger(__name__)
    logger.setLevel(logging.INFO)
    logging.basicConfig(format="%(levelname)s: %(message)s")
    logging.basicConfig(level=logging.INFO)

    input_dirs = [Path(input_dir) for input_dir in args.input]

    assert all([input_dir.is_dir() for input_dir in input_dirs]), "All input paths must be directories"

    separated_documents = {}
    seen_inventory_numbers = set()

    for input_dir in input_dirs:
        for sub_dir in input_dir.glob("*/"):
            inventory_number = sub_dir.name
            if inventory_number in seen_inventory_numbers:
                raise ValueError(f"Duplicate inventory number: {inventory_number}")
            seen_inventory_numbers.add(inventory_number)

            separated_documents[inventory_number] = {}
            previous_image_size = None
            for i, image_path in enumerate(get_file_paths(sub_dir, supported_image_formats, disable_check=True)):
                if i == 0:
                    image_size = imagesize.get(image_path)
                    current_document = image_path.name
                    separated_documents[inventory_number][current_document] = {
                        "numbers": [i + 1],
                        "sizes": [image_size],
                        "paths": [image_path],
                    }
                else:
                    image_size = imagesize.get(image_path)
                    if get_size_match(previous_image_size, image_size, 0.1):
                        separated_documents[inventory_number][current_document]["numbers"].append(i + 1)
                        separated_documents[inventory_number][current_document]["sizes"].append(image_size)
                        separated_documents[inventory_number][current_document]["paths"].append(image_path)
                    else:
                        current_document = image_path.name
                        separated_documents[inventory_number][current_document] = {
                            "numbers": [i + 1],
                            "sizes": [image_size],
                            "paths": [image_path],
                        }
                previous_image_size = image_size

    for inventory_number, documents in separated_documents.items():
        if len(documents) < 1:
            print(f"Inventory number {inventory_number} has no documents")
            continue

    total_documents = 0
    length_of_documents = Counter()

    for inventory_number, documents in separated_documents.items():
        total_documents += len(documents)
        length_of_documents += Counter([len(document["numbers"]) for document in documents.values()])

    logger.info(f"Total inventory numbers: {len(separated_documents)}")
    logger.info(
        f"Total scans: {sum(length*number_of_documents for length, number_of_documents in length_of_documents.items())}"
    )
    logger.info(f"Total documents: {total_documents}")
    logger.info(f"Document lengths: {OrderedDict(sorted(length_of_documents.items()))}")

    if not args.output:
        return

    if args.output_mode == "xlsx":
        assert args.output.endswith(".xlsx"), "Output file must be an xlsx file"

        workbook = Workbook()
        workbook.remove(workbook["Sheet"])

        main_sheet = workbook.create_sheet("Main")

        main_a_title = "Inventory number"
        main_b_title = "Dossier link"
        main_c_title = "Number of documents"

        main_sheet["A1"] = main_a_title
        main_sheet["B1"] = main_b_title
        main_sheet["C1"] = main_c_title

        main_a_width = len(main_a_title)
        main_b_width = len(main_b_title)
        main_c_width = len(main_c_title)

        for i, (inventory_number, documents) in enumerate(separated_documents.items(), start=2):
            main_sheet[f"A{i}"] = f'=HYPERLINK("#\'{inventory_number}\'!A1", "{inventory_number}")'
            spinque_link = f"https://cloud.spinque.com/oorlogvoorderechter/explore/dossier/{inventory_number}"
            main_sheet[f"B{i}"] = f'=HYPERLINK("{spinque_link}", "{spinque_link}")'
            main_sheet[f"C{i}"] = len(documents)

            main_a_width = max(main_a_width, len(str(inventory_number)))
            main_b_width = max(main_b_width, len(str(spinque_link)))
            main_c_width = max(main_c_width, len(str(len(documents))))

            inventory_sheet = workbook.create_sheet(inventory_number)
            inventory_a_title = "Start of document"
            inventory_b_title = "Scan name"
            inventory_c_title = "Number of pages"
            inventory_d_title = "Page numbers"
            inventory_sheet["A1"] = inventory_a_title
            inventory_sheet["B1"] = inventory_b_title
            inventory_sheet["C1"] = inventory_c_title
            inventory_sheet["D1"] = inventory_d_title
            length_a = len(inventory_a_title)
            length_b = len(inventory_b_title)
            length_c = len(inventory_c_title)
            length_d = len(inventory_d_title)

            for j, (document_name, document) in enumerate(documents.items(), start=2):
                spinque_link = (
                    f"https://cloud.spinque.com/oorlogvoorderechter/explore/dossier/{inventory_number}/{document['numbers'][0]}"
                )
                a = f'=HYPERLINK("{spinque_link}", "{spinque_link}")'
                length_a = max(length_a, len(str(spinque_link)))
                inventory_sheet[f"A{j}"] = a

                b = document_name
                length_b = max(length_b, len(b))
                inventory_sheet[f"B{j}"] = b

                c = len(document["numbers"])
                length_c = max(length_c, len(str(c)))
                inventory_sheet[f"C{j}"] = c

                d = ",".join(map(str, document["numbers"]))
                length_d = min(100, max(length_d, len(d)))
                inventory_sheet[f"D{j}"] = d

            inventory_sheet.column_dimensions["A"].width = length_a
            inventory_sheet.column_dimensions["B"].width = length_b
            inventory_sheet.column_dimensions["C"].width = length_c
            inventory_sheet.column_dimensions["D"].width = length_d

        main_sheet.column_dimensions["A"].width = main_a_width
        main_sheet.column_dimensions["B"].width = main_b_width
        main_sheet.column_dimensions["C"].width = main_c_width

        output_path = Path(args.output)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        workbook.save(output_path)
        logger.info(f"Separation ground truth saved to {output_path}")

    if args.output_mode == "dirs":
        output_dir = Path(args.output)
        output_dir.mkdir(parents=True, exist_ok=True)

        for inventory_number, documents in separated_documents.items():
            inventory_number_dir = output_dir.joinpath(inventory_number)
            inventory_number_dir.mkdir(parents=True, exist_ok=True)
            for document_name, document in documents.items():
                document_dir = inventory_number_dir.joinpath(document_name)
                document_dir.mkdir(parents=True, exist_ok=True)
                for image_path in document["paths"]:
                    copy_mode(image_path, document_dir / image_path.name)

        logger.info(f"Separation ground truth saved to {output_dir}")


if __name__ == "__main__":
    args = get_arguments()
    main(args)
